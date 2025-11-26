"""
Reports Export System - Excel and PDF
نظام تصدير التقارير
"""
import io
from datetime import datetime
from typing import List, Optional
from loguru import logger

# Excel export
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel export unavailable")

# PDF export
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False
    logger.warning("reportlab not installed - PDF export unavailable")

from database.models.user import User
from database.models.assignment import Assignment


class ReportGenerator:
    """Generate various reports"""
    
    @staticmethod
    async def generate_students_excel(course_id: Optional[str] = None) -> Optional[io.BytesIO]:
        """Generate Excel report of students"""
        if not EXCEL_AVAILABLE:
            logger.error("Excel export not available")
            return None
        
        try:
            # Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Students Report"
            
            # Headers
            headers = ['#', 'الاسم', 'البريد الإلكتروني', 'الهاتف', 'تاريخ التسجيل', 
                      'آخر نشاط', 'الدورات المسجلة', 'الواجبات المسلمة', 'المعدل']
            
            # Style headers
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Get students
            if course_id:
                users = await User.find().to_list()
                students = [u for u in users if u.has_approved_course(course_id)]
            else:
                students = await User.find().to_list()
            
            # Get all assignments
            assignments = await Assignment.find().to_list()
            
            # Fill data
            for idx, student in enumerate(students, 2):
                # Calculate stats
                enrolled_count = len([c for c in student.courses if c.approval_status == 'approved'])
                
                submitted_count = 0
                grades = []
                
                for assignment in assignments:
                    submission = assignment.get_submission(str(student.telegram_id))
                    if submission:
                        submitted_count += 1
                        if submission.grade is not None:
                            grades.append(submission.grade / assignment.max_grade * 100)
                
                avg_grade = sum(grades) / len(grades) if grades else 0
                
                # Write row
                row_data = [
                    idx - 1,
                    student.full_name,
                    student.email,
                    student.phone,
                    student.registered_at.strftime('%Y-%m-%d'),
                    student.last_active.strftime('%Y-%m-%d'),
                    enrolled_count,
                    submitted_count,
                    f"{avg_grade:.1f}%"
                ]
                
                for col, value in enumerate(row_data, 1):
                    cell = ws.cell(row=idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            # Save to BytesIO
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            logger.info(f"Excel report generated for {len(students)} students")
            return buffer
            
        except Exception as e:
            logger.error(f"Error generating Excel report: {e}")
            return None
    
    @staticmethod
    async def generate_grades_excel(course_id: str, assignment_id: Optional[str] = None) -> Optional[io.BytesIO]:
        """Generate Excel report of grades"""
        if not EXCEL_AVAILABLE:
            return None
        
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Grades Report"
            
            # Headers
            headers = ['#', 'الاسم', 'الواجب', 'الدرجة', 'النسبة', 'الحالة', 
                      'تاريخ التسليم', 'في الوقت المحدد']
            
            # Style headers
            header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            
            # Get data
            if assignment_id:
                assignments = [await Assignment.find_one(Assignment.id == assignment_id)]
            else:
                assignments = await Assignment.find(
                    Assignment.related_id == course_id
                ).to_list()
            
            row = 2
            for assignment in assignments:
                if not assignment:
                    continue
                
                for submission in assignment.submissions:
                    user = await User.find_one(User.telegram_id == int(submission.user_id))
                    if not user:
                        continue
                    
                    on_time = "نعم" if (assignment.deadline and submission.submitted_at <= assignment.deadline) else "لا"
                    status = "مصحح" if submission.status == "graded" else "قيد المراجعة"
                    percentage = f"{submission.grade / assignment.max_grade * 100:.1f}%" if submission.grade else "N/A"
                    
                    row_data = [
                        row - 1,
                        user.full_name,
                        assignment.title,
                        f"{submission.grade}/{assignment.max_grade}" if submission.grade else "N/A",
                        percentage,
                        status,
                        submission.submitted_at.strftime('%Y-%m-%d %H:%M'),
                        on_time
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        ws.cell(row=row, column=col, value=value)
                    
                    row += 1
            
            # Auto-adjust columns
            for col in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 15
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            return buffer
            
        except Exception as e:
            logger.error(f"Error generating grades Excel: {e}")
            return None
    
    @staticmethod
    async def generate_student_report_pdf(telegram_id: int) -> Optional[io.BytesIO]:
        """Generate PDF report for individual student"""
        if not PDF_AVAILABLE:
            logger.error("PDF export not available")
            return None
        
        try:
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            # Get student data
            user = await User.find_one(User.telegram_id == telegram_id)
            if not user:
                return None
            
            # Title
            title = Paragraph(f"<b>Student Report: {user.full_name}</b>", styles['Title'])
            elements.append(title)
            elements.append(Spacer(1, 20))
            
            # Student info table
            info_data = [
                ['Email:', user.email],
                ['Phone:', user.phone],
                ['Registered:', user.registered_at.strftime('%Y-%m-%d')],
                ['Last Active:', user.last_active.strftime('%Y-%m-%d')]
            ]
            
            info_table = Table(info_data, colWidths=[150, 350])
            info_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            elements.append(info_table)
            elements.append(Spacer(1, 20))
            
            # Grades section
            elements.append(Paragraph("<b>Grades Summary</b>", styles['Heading2']))
            elements.append(Spacer(1, 10))
            
            # Get grades
            assignments = await Assignment.find().to_list()
            grades_data = [['Assignment', 'Score', 'Percentage', 'Status']]
            
            for assignment in assignments:
                submission = assignment.get_submission(str(telegram_id))
                if submission and submission.grade is not None:
                    percentage = f"{submission.grade / assignment.max_grade * 100:.1f}%"
                    status = "Passed" if submission.grade >= assignment.pass_grade else "Failed"
                    grades_data.append([
                        assignment.title,
                        f"{submission.grade}/{assignment.max_grade}",
                        percentage,
                        status
                    ])
            
            if len(grades_data) > 1:
                grades_table = Table(grades_data, colWidths=[200, 80, 100, 80])
                grades_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                elements.append(grades_table)
            else:
                elements.append(Paragraph("No graded assignments yet.", styles['Normal']))
            
            # Build PDF
            doc.build(elements)
            buffer.seek(0)
            
            logger.info(f"PDF report generated for user {telegram_id}")
            return buffer
            
        except Exception as e:
            logger.error(f"Error generating PDF report: {e}")
            return None
